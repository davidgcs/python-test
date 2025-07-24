from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

# Cabecera estándar
header = ["Asignatura", "Título", "Autores", "Edición / Año", "Editorial", "Notas"]

# Copia-pega aquí todas las filas del archivo completo con las que empiezas, más las filas nuevas que necesitas.
# Incluye todas las filas del ejemplo que enviaste y añade después las asignaturas/fuentes que faltaban.
filas = [
    # -------------------- DATOS INICIALES (según tu ejemplo EN IMAGEN) -------------------
    ["Anatomía Funcional i Embriologia de l'Aparell Locomotor","Benninghoff y Drenckhahn, compendio de anatomía","Drenckhahn D, Waschke J, Negrete JH","2010","Médica Panamericana",""],
    ["","Feneis: nomenclatura anatómica ilustrada","Dauber W, Feneis H, Spitzer G","2007 (5a ed.)","Masson",""],
    ["","Gray Anatomía para estudiantes","Drake RL, Vogl W, Mitchell AWM","2015 (3ª ed.)","Elsevier",""],
    ["","Wolf-Heidegger’s atlas de anatomía","Köpf-Maier P, Wolf-Heidegger G","2000 (5ª ed.)","Marban","Disponible en inglés"],
    ["","Fundamentos de anatomía: con orientación clínica","Moore KL, Agur AMR, Dalley AF","2015 (5ª ed.)","Wolters Kluwer",""],
    ["","Atlas de anatomía humana","Netter FH","2015 (6ª ed.)","Elsevier Masson","Disponible en inglés"],
    ["","Atlas de anatomía: con correlación clínica","Platzer W","2008 (9ª ed.)","Médica Panamericana","3 volúmenes"],
    ["","Atlas de anatomía humana: estudio fotográfico","Rohen JW, Yokochi C, Lütjen-Drecoll E","2015 (8ª ed.)","Elsevier",""],
    ["","Prometheus: texto y atlas de anatomía","Schünke M, Schulte E, Schumacher U","2015 (3ª ed.)","Médica Panamericana","3 volúmenes"],
    ["","Atlas de anatomía humana de Sobotta","Paulsen F, Waschke J (eds.)","2012 (23ª ed.)","Elsevier","3 volúmenes"],
    ["Biología Celular y del Desarrollo","Molecular Biology of the Cell","Alberts B, et al.","2014 (6th ed.)","Garland Science","También en castellano"],
    ["","Biología Molecular de la Célula","Alberts B, et al.","2016 (6ª ed.)","Omega","También en inglés"],
    ["","Embriología humana y biología del desarrollo","Carlson BM","2019 (6ª ed.)","Elsevier",""],
    ["","Biología del Desarrollo","Gilbert SF","2019 (12ª ed.)","Médica Panamericana","También 11ª ed. en inglés"],
    ["","Embriología humana","Larsen WJ","2020 (6ª ed.)","Elsevier Science","También 5ª ed. en inglés"],
    ["","Embriología Clínica","Moore KL, Persaud TVN","2020 (11ª ed.)","Elsevier",""],
    ["","Embriología médica de Langman","Sadler TW","2019 (14ª ed.)","Lippincott Williams & Wilkins",""],
    ["","Principles of Development","Wolpert L, Tickle C, Martínez-Arias A","2019 (6th ed.)","Oxford University Press",""],
    ["Bioestadística, Epidemiología e Investigación","An Introduction to Medical Statistics","Bland M","2015 (4ª ed.)","Oxford",""],
    ["","Biostatistics","Daniel WW, Cross CL","2014 (10th ed.)","John Wiley & Sons","También 4ª ed. en castellano"],
    ["","Epidemiología","Gordis L","2014 (5ª ed.)","Elsevier","También en inglés"],
    ["","A Pocket Guide to Epidemiology","Kleinbaum DG, Sullivan KM, Barker ND","2007","Springer",""],
    ["","Fundamentals of Biostatistics","Rosner BA","2011 (7ª ed.)","Cengage Learning",""],

    # -------------------- FILAS NUEVAS (añade aquí el BLOQUE NUEVO) ---------------------

    # Biofísica Médica General
    ["Biofísica Médica General", "Física para las ciencias de la vida", "Cromer AH", "1981 (2ª ed.)", "Reverté", ""],
    ["", "Física en la ciencia y en la industria", "Cromer AH", "1986", "Reverté", ""],
    ["", "Física y biofísica: radiaciones", "Dutreix J, Desgrez A, Bok B", "1980", "AC", ""],
    ["", "Fisiología humana", "Fernández-Tresguerres JA... [et al.]", "2010 (4ª ed.)", "McGraw-Hill Interamericana Editores", ""],
    ["", "Biofísica", "Frumento AS", "1995 (3ª ed.)", "Mosby/Doyma Libros", ""],
    ["", "Física", "Kane JW, Sternheim MM", "1989 (2ª ed.)", "Reverté", ""],
    ["", "Física biológica: energía, información, vida", "Nelson PC", "2005", "Reverté", ""],
    ["", "Cellular biophysics", "Weiss TF", "1996", "The MIT Press", "2v."],
    ["", "Biofísica", "Yushimito Rubiños L", "2007", "Manual Moderno", ""],
    ["", "Física para la ciencia y la tecnología", "Tipler PA, Mosca G", "2010 (6a ed.)", "Reverté", ""],
    ["", "Física para ciencias de la vida", "David Jou Mirabent, Josep Enric Llebot Rabagliati, Carlos Pérez García", "2009 (2a ed.)", "McGraw-Hill/Interamericana de España", ""],
    ["", "Biological physics: energy, information, life", "Nelson PC, Marko Radosavljevic, Sarina Bromberg", "2014 (5th print.)", "W.H. Freeman", ""],

    # Biología Celular
    ["Biología Celular", "Biología molecular de la célula", "Alberts B, et al.", "2016 (6ª ed.)", "Omega", "También 6a ed. (2015) en inglés"],
    ["", "El Mundo de la célula", "Becker B, et al.", "2007 (6a edición)", "Addison Wesley", "También 9a ed. (2016) en inglés"],
    ["", "La célula", "Cooper GM, Hausman RM", "2017 (7a edición)", "Marbán", "También 7a ed. (2016) en inglés"],
    ["", "Biología celular y molecular", "Lodish H, et al.", "2016 (7a edición)", "Médica Panamericana", "También 8a ed. (2016) en inglés"],
    ["", "Cell biology", "Pollard TD, Earnshaw WC, Lippincott-Schwartz J, Johnson GT", "2016 (3th. ed.)", "Saunders Elsevier", ""],
    ["", "Physical biology of the cell", "Phillips R, Kondev J, Theriot J", "2013 (2nd ed.)", "Garland Science", ""],
    ["", "Molecular Cell Biology", "Lodish H, et al.", "2016 (8th edition)", "MacMillan Learning", ""],
    ["", "Molecular Biology of the Cell", "Alberts B, et al.", "2022 (7th edition)", "W.W. Norton", ""],
    ["", "Essential Cell Biology", "Alberts B, et al.", "2013 (4th edition)", "Garland Science", ""],

    # Biología Molecular
    ["Biología Molecular", "Biología molecular de la célula", "Alberts B, Wilson J, Hunt T", "2016 (6ª ed.)", "Omega", "También en inglés"],
    ["", "Bioquímica médica", "Baynes JW, Dominiczak MH", "2015 (4ª ed.)", "Elsevier", "También en inglés"],
    ["", "El mundo de la célula", "Becker WM, Kleinsmith LH, Hardin J", "2007 (6ª ed.)", "Addison Wesley", "También 9a ed. en inglés"],
    ["", "Biologia", "Campbell NA, Reece JB", "2007 (7ª ed.)", "Médica Panamericana", "También 10a ed. en inglés"],
    ["", "Bioquímica ilustrada: bioquímica y biología molecular en la era posgenómica", "Campbell PN, Smith AD, Peters TJ", "2006 (5ª ed.)", "Masson", ""],
    ["", "Bioquímica: libro de texto con aplicaciones clínicas", "Devlin TM", "2015 (4ª ed.)", "Reverté", "También 7a ed. en inglés"],
    ["", "Bioquímica", "Ferrier DR, Harvey RA", "2014 (6ª ed.)", "Wolters Kluwer Health/Lippincott Williams & Wilkins", "También 6a ed. en inglés"],
    ["", "Lewin’s genes XI", "Krebs JE, Goldstein ES, Kilpatrick ST", "2014", "Jones & Bartlett Learning", ""],
    ["", "Lewin’s essential genes", "Krebs JE, Goldstein ES, Kilpatrick ST", "2013 (3rd ed.)", "Jones and Bartlett Publishers", "También 2a ed. en castellà"],
    ["", "Molecular cell biology", "Lodish H, et al.", "2016 (8th. ed.)", "Freeman", ""],
    ["", "Bioquímica médica básica: un enfoque clínico", "Lieberman M, Marks A, Peet A", "2013 (4a ed.)", "Wolters Kluwer/Lippincott Williams & Wilkins", "También 4a ed. en inglés"],
    ["", "Bioquímica", "Mathews CHK, et al.", "2013 (4ª ed.)", "Pearson Educación", "También 4a ed. en inglés"],
    ["", "Bioquímica: las bases moleculares de la vida", "Mckee T, Mckee JR", "2014 (5ª ed.)", "McGraw-Hill/Interamericana", ""],
    ["", "Bioquímica con aplicaciones clínicas", "Stryer L, Berg JM, Tymoczko JL", "2015 (7ª ed.)", "Reverté", "También 8a ed. en inglés"],
    ["", "Bioquímica", "Voet D, Voet JG", "2006 (3ª ed.)", "Médica panamericana", "También 4a ed. en inglés"],
    ["", "Molecular biology of the gene", "Watson JD, et al.", "2013 (7th ed.)", "Pearson", "También 7a ed. en castellà"],
    ["", "Biochemistry", "Jeremy M. Berg, John L. Tymoczko, Gregory J. Gatto, Jr., Lubert Stryer", "2015 (8th edition)", "", ""],
    ["", "NCBI Home Page", "", "", "", "Bases de dades"],
    ["", "PDB lite", "", "", "", "Bases de dades"],
    ["", "The JASPAR Database", "", "", "", "Bases de dades per analitzar regions reguladores gèniques"],
    ["", "UNIPROT", "", "", "", "Bases de dades"],
    ["", "The Medical Biochemistry Page", "", "", "", "Pàgina web"],
    ["", "http://www.ncbi.nlm.nih.gov/BLAST/", "", "", "", "Pàgina web"],
    ["", "GTEX Portal", "", "", "", "https://gtexportal.org/home/"],
    ["", "Protter", "", "", "", "http://wlab.ethz.ch/protter/start/"],

    # Bioquímica Básica
    ["Bioquímica Básica", "", "", "", "", "Nada aún"],

    # Histología Humana
    ["Histología Humana", "Geneser histología : 4a. edición", "Brüel A, Geneser F", "2015", "Editorial Médica Panamericana", ""],
    ["", "Histología : atlas en color y texto", "Gartner LP", "2018 (7a ed.)", "Wolters Kluwer", "También 7a ed. (2018) en inglés"],
    ["", "Texto de histología : atlas a color", "Gartner LP", "2017 (7a ed.)", "Elsevier", "También 4a ed. (2017) en inglés"],
    ["", "Histología básica : texto y atlas", "Junqueira LCU, Carneiro J", "2015 (12a ed.)", "Editorial Médica Panamericana", "También 14a ed. en inglés"],
    ["", "Histology and cell biology : an introduction to pathology", "Kierszenbaum AL, Tres LL", "2016 (4th ed.)", "Elsevier/Saunders", "También 4a ed. (2016) en castellano"],
    ["", "Atlas color de citología e histología", "Kühnel W", "2005 (11a ed. corr. y)", "Medica Panamericana", "También 4th ed. (2003) en inglés."],
    ["", "Stevens y Lowe histología humana: cuarta edición", "Lowe JS, Anderson PG", "2015", "Elsevier", "También 4a ed. en inglés"],
    ["", "Histology: a text and atlas : with correlated cell and molecular biology", "Ross MH, Pawlina W", "2016 (7th ed.)", "Wolters Kluwer Health", "También 7a ed. en castellano"],
    ["", "Wheater : histología funcional : texto y atlas en color", "Young B, O’Dowd G, Woodford P", "2014 (6a ed.)", "Elsevier", "También 6a ed. en inglés"],
    ["", "http://www.udel.edu/biology/Wags/histopage/histopage.htm", "", "", "", "Pàgina web"],
    ["", "http://www.kumc.edu/instruction/medicine/anatomy/histoweb/index.htm", "", "", "", "Pàgina web"],
    ["", "http://www.lumen.luc.edu/lumen/MedEd/Histo/frames/histo_frames.html", "", "", "", "Pàgina web"],
    ["", "http://cal.vet.upenn.edu/index.php?page=histology", "", "", "", "Pàgina web"],
    ["", "https://ubmedicina.ventana-vector.com/", "", "", "", "Pàgina web"],

    # Introducción a la Salud, Antropología, Demografía. Historia de la Medicina
    ["Introducción a la Salud, Antropología, Demografía. Historia de la Medicina", "¿Me está escuchando doctor? : un viaje por la mente de los médicos", "Groopman JE", "2008", "RBA", ""],
    ["", "Breve historia de la medicina", "López Piñero JM", "2000", "Alianza", ""],
    ["", "Semmelweis", "Louis-Ferdinand Celine", "2014", "Ed Marbot", ""],

    # Inglés Médico
    ["Inglés Médico", "Text electrònic — Articles actuals rellevants i recursos multimèdia — Materials publicats al Campus Virtual", "", "", "", ""]
]

# -- Crea el archivo Excel --
wb = Workbook()
ws = wb.active
ws.title = "Sheet1"
ws.append(header)

# Inserta todas las filas
for fila in filas:
    ws.append(fila)

# -- Ajusta anchuras de columna (editables) --
ancho_cols = [28, 52, 36, 26, 35, 35]
for i, ancho in enumerate(ancho_cols, 1):
    ws.column_dimensions[get_column_letter(i)].width = ancho

# -- Combina celdas de la columna Asignatura con mismo valor (como en tu Excel) --
asig_col = 1
asig_map = defaultdict(list)
current_asig = None
for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=1, values_only=True), start=2):
    val = row[0]
    if val:
        current_asig = val
    asig_map[current_asig].append(idx)
for fila_indices in asig_map.values():
    if len(fila_indices) > 1:
        ws.merge_cells(start_row=fila_indices[0], start_column=asig_col,
                       end_row=fila_indices[-1], end_column=asig_col)

# --- Estilos: encabezado, filas alternas, wrap, bordes ---
header_fill = PatternFill("solid", fgColor="1976D2")
header_font = Font(bold=True, color="FFFFFF")
alt_1 = PatternFill("solid", fgColor="E3F2FD")
alt_2 = PatternFill("solid", fgColor="FFFFFF")
border = Border(
    left=Side(style='thin', color='888888'),
    right=Side(style='thin', color='888888'),
    top=Side(style='thin', color='888888'),
    bottom=Side(style='thin', color='888888'),
)
wrap = Alignment(vertical="top", horizontal="left", wrap_text=True)

for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row), start=1):
    if idx == 1:
        for cell in row:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
    else:
        fill = alt_1 if idx % 2 == 0 else alt_2
        for cell in row:
            cell.fill = fill
            cell.border = border
            cell.alignment = wrap
            cell.font = Font(name="Calibri", size=11)

# Ajusta altura automática por número de líneas estimadas
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    max_lines = max(str(cell.value).count('\n')+1 if cell.value else 1 for cell in row)
    ws.row_dimensions[row[0].row].height = max(18, 16*max_lines)

# -- Guarda Excel --
nombre_archivo = "Asignaturas_Medicina_Completo_Estilo.xlsx"
wb.save(nombre_archivo)
print(f"\nArchivo creado: {nombre_archivo}\n¡Ábrelo en Excel para ver los estilos y la agrupación por asignatura!")
